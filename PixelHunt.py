#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║  ██████╗ ██╗██╗  ██╗███████╗██╗     ██╗  ██╗██╗   ██╗███╗   ██╗████████╗   ║
║  ██╔══██╗██║╚██╗██╔╝██╔════╝██║     ██║  ██║██║   ██║████╗  ██║╚══██╔══╝   ║
║  ██████╔╝██║ ╚███╔╝ █████╗  ██║     ███████║██║   ██║██╔██╗ ██║   ██║      ║
║  ██╔═══╝ ██║ ██╔██╗ ██╔══╝  ██║     ██╔══██║██║   ██║██║╚██╗██║   ██║      ║
║  ██║     ██║██╔╝ ██╗███████╗███████╗██║  ██║╚██████╔╝██║ ╚████║   ██║      ║
║  ╚═╝     ╚═╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝  ╚═╝ ╚═════╝ ╚═╝  ╚═══╝   ╚═╝      ║
║                                                                               ║
║                  🎯  Image Object Detection Pipeline                          ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import argparse
import cv2
import pandas as pd
from pathlib import Path
from datetime import datetime
from ultralytics import YOLO
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, BarColumn,
    TextColumn, TimeElapsedColumn, MofNCompleteColumn,
)
from rich.text import Text
from rich.align import Align
from rich.rule import Rule
from rich import box
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ✅ استيراد المترجم
from translations import ClassTranslator

# ════════════════════════════════════════════════
#   GLOBALS
# ════════════════════════════════════════════════

console = Console()

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif"}

REPORT_RTL = True

BANNER = r"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║  ██████╗ ██╗██╗  ██╗███████╗██╗     ██╗  ██╗██╗   ██╗███╗   ██╗████████╗   ║
║  ██╔══██╗██║╚██╗██╔╝██╔════╝██║     ██║  ██║██║   ██║████╗  ██║╚══██╔══╝   ║
║  ██████╔╝██║ ╚███╔╝ █████╗  ██║     ███████║██║   ██║██╔██╗ ██║   ██║      ║
║  ██╔═══╝ ██║ ██╔██╗ ██╔══╝  ██║     ██╔══██║██║   ██║██║╚██╗██║   ██║      ║
║  ██║     ██║██╔╝ ██╗███████╗███████╗██║  ██║╚██████╔╝██║ ╚████║   ██║      ║
║  ╚═╝     ╚═╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝  ╚═╝ ╚═════╝ ╚═╝  ╚═══╝   ╚═╝      ║
║                                                                               ║
║                  🎯  Image Object Detection Pipeline                          ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

CLASS_COLORS = [
    (255,  56,  56), (255, 157, 151), (255, 112,  31),
    (255, 178,  29), (207, 210,  49), ( 72, 249,  10),
    (146, 204,  23), ( 61, 219, 134), ( 26, 147,  52),
    (  0, 212, 187), ( 44, 153, 168), (  0, 194, 255),
    ( 52,  69, 147), (100, 115, 255), (  0,  24, 236),
    (132,  56, 255), ( 82,   0, 133), (203,  56, 255),
    (255, 149, 200), (255,  55, 199),
]

def get_color(class_id: int) -> tuple:
    return CLASS_COLORS[class_id % len(CLASS_COLORS)]


# ════════════════════════════════════════════════
#   EXCEL STYLING
# ════════════════════════════════════════════════

def _apply_sheet_style(ws, is_rtl: bool = True):
    ws.sheet_view.rightToLeft = is_rtl
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="EEF2FF")
    h_align     = "right" if is_rtl else "left"

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal=h_align, vertical="center")

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"


# ════════════════════════════════════════════════
#   LOAD IMAGES
# ════════════════════════════════════════════════

def load_images(folder: Path) -> list[Path]:
    images = sorted([
        f for f in folder.iterdir()
        if f.suffix.lower() in IMAGE_EXTENSIONS
    ])
    if not images:
        console.print(Panel(
            "[bold red]❌  No images found in the specified folder![/bold red]\n"
            f"[dim]Path: {folder.resolve()}[/dim]",
            title="[bold red]Error[/bold red]",
            border_style="red",
        ))
        sys.exit(1)
    return images


# ════════════════════════════════════════════════
#   CORE DETECTION
# ════════════════════════════════════════════════

def run_detection(
    model      : YOLO,
    images     : list[Path],
    conf       : float,
    translator : ClassTranslator,          # ✅ مضاف
) -> list[dict]:
    all_results = []

    with Progress(
        SpinnerColumn(style="cyan"),
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(bar_width=38, style="cyan", complete_style="bright_cyan"),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=False,
    ) as progress:
        task = progress.add_task("🔍  Detecting objects...", total=len(images))

        for img_path in images:
            results    = model(str(img_path), conf=conf, verbose=False)[0]
            img        = cv2.imread(str(img_path))
            h, w       = img.shape[:2]
            ts         = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            detections = []

            for b in results.boxes:
                cls_id     = int(b.cls[0])
                cls_name   = model.names[cls_id]
                # ✅ الاسم المعروض: "سيارة (car)" أو "car"
                cls_display = translator.translate_display(cls_name)
                confidence = float(b.conf[0])
                x1, y1, x2, y2 = map(int, b.xyxy[0])

                detections.append({
                    "image_name"  : img_path.name,
                    "image_path"  : str(img_path.resolve()),
                    "class_id"    : cls_id,
                    "class_name"  : cls_display,        # ✅ ثنائي اللغة
                    "class_en"    : cls_name,            # ✅ الإنجليزي للمعالجة الداخلية
                    "confidence"  : round(confidence, 4),
                    "x1"          : x1, "y1": y1,
                    "x2"          : x2, "y2": y2,
                    "obj_width"   : x2 - x1,
                    "obj_height"  : y2 - y1,
                    "img_width"   : w,
                    "img_height"  : h,
                    "detected_at" : ts,
                })

            if not detections:
                detections.append({
                    "image_name"  : img_path.name,
                    "image_path"  : str(img_path.resolve()),
                    "class_id"    : None,
                    "class_name"  : "no_detection",
                    "class_en"    : "no_detection",
                    "confidence"  : 0.0,
                    "x1": None, "y1": None,
                    "x2": None, "y2": None,
                    "obj_width"   : None, "obj_height": None,
                    "img_width"   : w,
                    "img_height"  : h,
                    "detected_at" : ts,
                })

            all_results.extend(detections)
            progress.advance(task)

    return all_results


# ════════════════════════════════════════════════
#   MODE 1 — REPORT
# ════════════════════════════════════════════════

def _save_excel(df: pd.DataFrame, path: Path, is_rtl: bool):
    detected = df[df["class_name"] != "no_detection"]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ── Sheet 1: Raw Detections ──
        df.to_excel(writer, sheet_name="Raw Detections", index=False)

        # ── Sheet 2: Class Summary ──
        if not detected.empty:
            summary = (
                detected.groupby("class_name")
                .agg(
                    total_detections = ("class_name", "count"),
                    images_count     = ("image_name", "nunique"),
                    avg_confidence   = ("confidence", "mean"),
                    min_confidence   = ("confidence", "min"),
                    max_confidence   = ("confidence", "max"),
                    avg_width        = ("obj_width",  "mean"),
                    avg_height       = ("obj_height", "mean"),
                )
                .round(4)
                .reset_index()
                .sort_values("total_detections", ascending=False)
            )
            summary.to_excel(writer, sheet_name="Class Summary", index=False)

        # ── Sheet 3: Per Image Summary ──
        if not detected.empty:
            img_summary = (
                detected.groupby("image_name")
                .agg(
                    total_objects  = ("class_name", "count"),
                    unique_classes = ("class_name", "nunique"),
                    classes_found  = ("class_name", lambda x: ", ".join(sorted(x.unique()))),
                    avg_confidence = ("confidence", "mean"),
                    img_width      = ("img_width",  "first"),
                    img_height     = ("img_height", "first"),
                )
                .round(4)
                .reset_index()
                .sort_values("total_objects", ascending=False)
            )
            img_summary.to_excel(writer, sheet_name="Per Image Summary", index=False)

        # ── Sheet 4: Stats Overview ──
        stats = {
            "Metric": [
                "Total Images Scanned", "Images With Detections",
                "Images Without Detections", "Total Objects Detected",
                "Unique Classes Found", "Average Confidence",
                "Highest Confidence", "Lowest Confidence", "Generated At",
            ],
            "Value": [
                df["image_name"].nunique(),
                detected["image_name"].nunique() if not detected.empty else 0,
                df[df["class_name"] == "no_detection"]["image_name"].nunique(),
                len(detected),
                detected["class_name"].nunique() if not detected.empty else 0,
                f"{detected['confidence'].mean():.2%}" if not detected.empty else "N/A",
                f"{detected['confidence'].max():.2%}" if not detected.empty else "N/A",
                f"{detected['confidence'].min():.2%}" if not detected.empty else "N/A",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        }
        pd.DataFrame(stats).to_excel(writer, sheet_name="Stats Overview", index=False)

        for sheet_name in writer.sheets:
            _apply_sheet_style(writer.sheets[sheet_name], is_rtl=is_rtl)


def _print_report_tables(df: pd.DataFrame):
    detected   = df[df["class_name"] != "no_detection"]
    total_imgs = df["image_name"].nunique()
    total_objs = len(detected)
    total_cls  = detected["class_name"].nunique() if not detected.empty else 0
    avg_conf   = detected["confidence"].mean()    if not detected.empty else 0.0

    console.print()
    console.print(Panel(
        Align.center(Text.assemble(
            ("  🎯  PixelHunt — Detection Report  \n\n", "bold bright_white"),
            ("  📸  Images Scanned   ", "dim white"), (f"{total_imgs:>6}\n",  "bold cyan"),
            ("  📦  Objects Found    ", "dim white"), (f"{total_objs:>6}\n",  "bold yellow"),
            ("  🏷️   Unique Classes   ", "dim white"), (f"{total_cls:>6}\n",  "bold magenta"),
            ("  📈  Avg Confidence   ", "dim white"), (f"{avg_conf:>5.1%}\n", "bold green"),
        )),
        border_style="bright_blue",
        padding=(1, 6),
    ))

    if detected.empty:
        console.print("[bold red]  ⚠️  No objects detected in any image.[/bold red]")
        return

    # ── جدول 1: ملخص الكلاسات ──
    console.print()
    console.rule("[bold cyan]📊  Class Summary[/bold cyan]")
    console.print()

    tbl1 = Table(
        title="🏷️  Detected Classes",
        caption=f"Total: {total_objs} objects  •  {total_cls} classes",
        box=box.DOUBLE_EDGE, border_style="bright_blue",
        header_style="bold bright_white on navy_blue",
        show_lines=True, padding=(0, 1),
        title_style="bold cyan", caption_style="dim cyan",
    )
    tbl1.add_column("Rank",         style="bold white",   justify="center")
    tbl1.add_column("Class",        style="bold yellow",  justify="left")
    tbl1.add_column("Objects",      style="bold white",   justify="center")
    tbl1.add_column("Images",       style="cyan",         justify="center")
    tbl1.add_column("Avg Conf",     style="green",        justify="center")
    tbl1.add_column("Max Conf",     style="bright_green", justify="center")
    tbl1.add_column("Distribution", style="cyan",         justify="left")

    class_summary = (
        detected.groupby("class_name")
        .agg(
            count    = ("class_name", "count"),
            imgs     = ("image_name", "nunique"),
            avg_conf = ("confidence", "mean"),
            max_conf = ("confidence", "max"),
        )
        .sort_values("count", ascending=False)
    )

    for rank, (cls, row) in enumerate(class_summary.iterrows(), 1):
        medal   = {1: "🥇", 2: "🥈", 3: "🥉"}.get(rank, f"  {rank}.")
        bar_len = int((row["count"] / class_summary["count"].max()) * 14)
        bar     = f"[cyan]{'█' * bar_len}[/cyan][dim]{'░' * (14 - bar_len)}[/dim]"
        tbl1.add_row(
            medal, str(cls),
            str(int(row["count"])), str(int(row["imgs"])),
            f"{row['avg_conf']:.1%}", f"{row['max_conf']:.1%}", bar,
        )

    console.print(Align.center(tbl1))

    # ── جدول 2: ملخص الصور ──
    console.print()
    console.rule("[bold yellow]🖼️   Per Image Summary[/bold yellow]")
    console.print()

    tbl2 = Table(
        title="📸  Image Results",
        caption=f"Scanned {total_imgs} images",
        box=box.DOUBLE_EDGE, border_style="bright_blue",
        header_style="bold bright_white on navy_blue",
        show_lines=True, padding=(0, 1),
        title_style="bold yellow", caption_style="dim cyan",
    )
    tbl2.add_column("Image Name", style="bold white",  justify="left",   no_wrap=True)
    tbl2.add_column("Objects",    style="bold yellow", justify="center")
    tbl2.add_column("Classes",    style="cyan",        justify="center")
    tbl2.add_column("Found",      style="dim white",   justify="left")
    tbl2.add_column("Avg Conf",   style="green",       justify="center")

    img_summary = (
        detected.groupby("image_name")
        .agg(
            objs    = ("class_name", "count"),
            classes = ("class_name", "nunique"),
            found   = ("class_name", lambda x: ", ".join(sorted(x.unique()))),
            conf    = ("confidence", "mean"),
        )
        .sort_values("objs", ascending=False)
    )

    for img, row in img_summary.iterrows():
        status = "🟢" if row["objs"] >= 5 else "🟡" if row["objs"] >= 2 else "🔵"
        found  = row["found"][:45] + ("…" if len(row["found"]) > 45 else "")
        tbl2.add_row(
            f"{status}  {img}",
            str(int(row["objs"])), str(int(row["classes"])),
            found, f"{row['conf']:.1%}",
        )

    console.print(Align.center(tbl2))


def mode_report(
    data       : list[dict],
    output_dir : Path,
    fmt        : str,
    is_rtl     : bool,
):
    report_dir = output_dir / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    df  = pd.DataFrame(data)

    saved_files = []

    if fmt in ("csv", "both"):
        path = report_dir / f"pixelhunt_report_{ts}.csv"
        df.to_csv(path, index=False, encoding="utf-8-sig")
        saved_files.append(("CSV", path))

    if fmt in ("excel", "both"):
        path = report_dir / f"pixelhunt_report_{ts}.xlsx"
        _save_excel(df, path, is_rtl=is_rtl)
        saved_files.append(("Excel", path))

    _print_report_tables(df)

    console.print()
    console.rule("[bold green]💾  Saved Files[/bold green]")
    for ftype, fpath in saved_files:
        console.print(f"  [bold green]✅  {ftype}[/bold green]  →  [cyan]{fpath}[/cyan]")
    console.print()


# ════════════════════════════════════════════════
#   MODE 2 — ANNOTATE
# ════════════════════════════════════════════════

def mode_annotate(
    model      : YOLO,
    images     : list[Path],
    output_dir : Path,
    conf       : float,
    translator : ClassTranslator,          # ✅ مضاف
):
    ann_dir = output_dir / "annotated"
    ann_dir.mkdir(parents=True, exist_ok=True)
    saved = 0

    with Progress(
        SpinnerColumn(style="yellow"),
        TextColumn("[bold yellow]{task.description}"),
        BarColumn(bar_width=38, style="yellow", complete_style="bright_yellow"),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("🎨  Annotating images...", total=len(images))

        for img_path in images:
            img     = cv2.imread(str(img_path))
            results = model(str(img_path), conf=conf, verbose=False)[0]

            for b in results.boxes:
                cls_id   = int(b.cls[0])
                cls_name = model.names[cls_id]
                conf_val = float(b.conf[0])
                x1, y1, x2, y2 = map(int, b.xyxy[0])
                color = get_color(cls_id)

                # ✅ الليبل ثنائي اللغة: "سيارة (car)"
                label = f"  {translator.translate_display(cls_name)}  {conf_val:.0%}  "

                cv2.rectangle(img, (x1, y1), (x2, y2), color, 3)
                cv2.rectangle(img, (x1+1, y1+1), (x2-1, y2-1), (255, 255, 255), 1)

                (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                cv2.rectangle(img, (x1, y1 - th - 14), (x1 + tw + 4, y1), color, -1)
                cv2.putText(img, label, (x1 + 2, y1 - 6),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            cv2.imwrite(str(ann_dir / img_path.name), img)
            saved += 1
            progress.advance(task)

    console.print()
    console.print(Panel(
        f"[bold green]✅  {saved} annotated images saved[/bold green]\n"
        f"[dim cyan]📁  {ann_dir.resolve()}[/dim cyan]",
        title="[bold white]🎨  Mode 2 — Complete[/bold white]",
        border_style="green",
        padding=(1, 4),
    ))


# ════════════════════════════════════════════════
#   MODE 3 — FILTER BY CLASS
# ════════════════════════════════════════════════

def mode_filter(
    model          : YOLO,
    images         : list[Path],
    output_dir     : Path,
    raw_tokens     : list[str],            # ✅ tokens خام من CLI
    conf           : float,
    is_rtl         : bool,
    translator     : ClassTranslator,      # ✅ مضاف
):
    # ── ① حل مشكلة الكلمتين + Validation ──
    valid_en, invalid = translator.match_classes(raw_tokens)

    # ⚠️ تحذير للكلاسات الخاطئة
    if invalid:
        console.print(Panel(
            f"[bold yellow]⚠️  الكلاسات التالية غير موجودة:[/bold yellow]  "
            f"[red]{', '.join(invalid)}[/red]\n\n"
            f"[dim]الكلاسات المتاحة:\n{translator.available_classes_display()}[/dim]",
            border_style="yellow",
            padding=(1, 3),
        ))

    # ❌ لو كل الكلاسات خاطئة
    if not valid_en:
        console.print("[bold red]❌  لا توجد كلاسات صحيحة للمتابعة![/bold red]")
        sys.exit(1)

    # ✅ عرض ما سيتم البحث عنه
    display_names = [translator.translate_display(c) for c in valid_en]
    console.print(
        f"[bold green]✅  سيتم البحث عن:[/bold green]  "
        f"[cyan]{', '.join(display_names)}[/cyan]\n"
    )

    target_lower = {c.lower() for c in valid_en}
    folder_name  = "filtered__" + "_".join(valid_en)
    filtered_dir = output_dir / folder_name
    filtered_dir.mkdir(parents=True, exist_ok=True)

    match_data   = []
    matched_imgs = []

    with Progress(
        SpinnerColumn(style="magenta"),
        TextColumn("[bold magenta]{task.description}"),
        BarColumn(bar_width=38, style="magenta", complete_style="bright_magenta"),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("🔎  Filtering by class...", total=len(images))

        for img_path in images:
            img     = cv2.imread(str(img_path))
            results = model(str(img_path), conf=conf, verbose=False)[0]

            hits = []
            for b in results.boxes:
                cls_id   = int(b.cls[0])
                cls_name = model.names[cls_id]
                conf_val = float(b.conf[0])
                if cls_name.lower() in target_lower:
                    hits.append((cls_id, cls_name, conf_val, b))

            if hits:
                matched_imgs.append(img_path.name)
                for cls_id, cls_name, conf_val, b in hits:
                    x1, y1, x2, y2 = map(int, b.xyxy[0])
                    color = get_color(cls_id)

                    # ✅ الليبل ثنائي اللغة
                    label = f"  ★ {translator.translate_display(cls_name)}  {conf_val:.0%}  "

                    cv2.rectangle(img, (x1, y1), (x2, y2), color, 3)
                    cv2.rectangle(img, (x1+1, y1+1), (x2-1, y2-1), (255, 255, 255), 1)

                    (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
                    cv2.rectangle(img, (x1, y1 - th - 14), (x1 + tw + 4, y1), color, -1)
                    cv2.putText(img, label, (x1 + 2, y1 - 6),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)

                    match_data.append({
                        "image_name"  : img_path.name,
                        "class_name"  : translator.translate_display(cls_name),  # ✅
                        "class_en"    : cls_name,
                        "confidence"  : round(conf_val, 4),
                        "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                        "obj_width"   : x2 - x1,
                        "obj_height"  : y2 - y1,
                        "detected_at" : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })

                cv2.imwrite(str(filtered_dir / img_path.name), img)

            progress.advance(task)

    # ── طباعة النتائج ──
    console.print()

    if not matched_imgs:
        console.print(Panel(
            f"[bold red]⚠️  No images found containing:[/bold red]\n"
            f"[yellow]  {', '.join(display_names)}[/yellow]",
            border_style="red", padding=(1, 4),
        ))
        return

    console.rule("[bold magenta]🎯  Filter Results[/bold magenta]")
    console.print()

    tbl = Table(
        title=f"🎯  Matched — {', '.join(display_names)}",
        caption=f"{len(matched_imgs)} matching images  •  {len(match_data)} total detections",
        box=box.DOUBLE_EDGE, border_style="bright_blue",
        header_style="bold bright_white on navy_blue",
        show_lines=True, padding=(0, 1),
        title_style="bold magenta", caption_style="dim cyan",
    )
    tbl.add_column("Image",      style="bold white",  justify="left",   no_wrap=True)
    tbl.add_column("Class",      style="bold yellow", justify="center")
    tbl.add_column("Confidence", style="bold green",  justify="center")
    tbl.add_column("Position",   style="cyan",        justify="center")
    tbl.add_column("Size (px)",  style="dim white",   justify="center")

    for d in match_data:
        tbl.add_row(
            d["image_name"],
            f"★  {d['class_name']}",
            f"{d['confidence']:.1%}",
            f"({d['x1']}, {d['y1']})",
            f"{d['obj_width']} × {d['obj_height']}",
        )

    console.print(Align.center(tbl))

    if match_data:
        report_dir = output_dir / "reports"
        report_dir.mkdir(exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = report_dir / f"pixelhunt_filter_{'_'.join(valid_en)}_{ts}.csv"
        pd.DataFrame(match_data).to_csv(csv_path, index=False, encoding="utf-8-sig")

        console.print()
        console.print(Panel(
            f"[bold green]✅  {len(matched_imgs)} images matched & saved[/bold green]\n"
            f"[dim cyan]📁  Folder  →  {filtered_dir.resolve()}[/dim cyan]\n"
            f"[dim cyan]📄  CSV     →  {csv_path}[/dim cyan]",
            title="[bold white]🎯  Mode 3 — Complete[/bold white]",
            border_style="magenta", padding=(1, 4),
        ))


# ════════════════════════════════════════════════
#   MAIN
# ════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        prog="pixelhunt",
        description="🎯 PixelHunt — Image Object Detection Pipeline",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--images",  "-i", default="images",
                        help="Images folder       (default: images/)")
    parser.add_argument("--output",  "-o", default="output",
                        help="Output folder        (default: output/)")
    parser.add_argument("--model",   "-m", default="yolo11n.pt",
                        help="YOLO model           (default: yolo11n.pt)")
    parser.add_argument("--mode",         type=int, choices=[1, 2, 3], required=True,
                        help="1 = Report\n2 = Annotate\n3 = Filter by class")
    parser.add_argument("--classes", "-c", nargs="+",
                        help=(
                            "Target classes for mode 3\n"
                            "  English : --classes car person\n"
                            "  Arabic  : --classes سيارة شخص\n"
                            "  Multi-word: --classes لوحة مفاتيح سيارة"
                        ))
    parser.add_argument("--format",  "-f", choices=["csv", "excel", "both"], default="both",
                        help="Report format for mode 1  (default: both)")
    parser.add_argument("--conf",         type=float, default=0.25,
                        help="Confidence threshold      (default: 0.25)")
    parser.add_argument("--dir",          choices=["LTR", "RTL"], default="RTL",
                        help="Report/Excel direction    (default: RTL)")
    # ✅ وسيط اللغة
    parser.add_argument("--lang",         default="en",
                        choices=ClassTranslator.available_languages(),
                        help="Output language  (default: en)\n  e.g. --lang ar")

    args   = parser.parse_args()
    is_rtl = (args.dir == "RTL")

    # ✅ تهيئة المترجم مرة واحدة وتمريره لكل الدوال
    translator = ClassTranslator(lang=args.lang)

    mode_label = {1: "Report", 2: "Annotate", 3: "Filter"}[args.mode]

    console.print()
    console.print(Panel(
        Align.center(Text.assemble(
            (BANNER, "bold cyan"),
            (f"\n  Mode  ", "dim white"),
            (f"{mode_label:^10}", "bold yellow"),
            (f"  |  Model  ", "dim white"),
            (f"{args.model:^16}", "bold green"),
            (f"  |  Lang  ", "dim white"),
            (f"{args.lang.upper()}  |  Dir  ", "bold magenta"),
            (f"{args.dir}\n", "bold magenta"),
        )),
        border_style="bright_blue",
        padding=(0, 2),
    ))
    console.print()

    images_dir = Path(args.images)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    console.print(f"[bold cyan]📦  Loading model:[/bold cyan]  [yellow]{args.model}[/yellow]")
    model = YOLO(args.model)

    images = load_images(images_dir)
    console.print(f"[bold cyan]🖼️   Images found:[/bold cyan]   [yellow]{len(images)}[/yellow]\n")

    if args.mode == 1:
        console.rule("[bold cyan]Mode 1 — Detection Report[/bold cyan]")
        # ✅ تمرير المترجم
        data = run_detection(model, images, args.conf, translator)
        mode_report(data, output_dir, args.format, is_rtl=is_rtl)

    elif args.mode == 2:
        console.rule("[bold yellow]Mode 2 — Annotate Images[/bold yellow]")
        # ✅ تمرير المترجم
        mode_annotate(model, images, output_dir, args.conf, translator)

    elif args.mode == 3:
        if not args.classes:
            console.print(
                "[bold red]❌  Mode 3 requires --classes\n"
                "    e.g.  python pixelhunt.py --mode 3 --lang ar --classes سيارة شخص[/bold red]"
            )
            sys.exit(1)
        console.rule("[bold magenta]Mode 3 — Filter by Class[/bold magenta]")
        # ✅ تمرير raw_tokens والمترجم
        mode_filter(model, images, output_dir, args.classes, args.conf, is_rtl, translator)

    console.print()
    console.rule("[bold green]✅  PixelHunt — Done[/bold green]")
    console.print(
        f"\n  [dim]Output saved to:[/dim]  "
        f"[bold cyan]{output_dir.resolve()}[/bold cyan]\n"
    )


if __name__ == "__main__":
    main()
