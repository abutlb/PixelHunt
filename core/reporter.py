# core/reporter.py

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  JSON — المخرج الافتراضي
# ─────────────────────────────────────────────

def save_json(results: list[dict], path: Path) -> Path:
    """
    يحفظ النتائج كـ JSON منظم.
    الشكل:
    {
        "generated_at": "...",
        "total_images": N,
        "total_objects": N,
        "results": [ { ...per image... } ]
    }
    """
    total_objects = sum(r["total_objects"] for r in results)

    output = {
        "generated_at" : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_images" : len(results),
        "total_objects": total_objects,
        "results"      : results,
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(output, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


# ─────────────────────────────────────────────
#  تحويل النتائج لـ DataFrame مسطح
# ─────────────────────────────────────────────

def results_to_df(results: list[dict]) -> pd.DataFrame:
    """يحول list of dicts المتداخلة إلى DataFrame مسطح."""
    rows = []
    for r in results:
        if not r["detections"]:
            rows.append({
                "image_name"   : r["image_name"],
                "source_url"   : r.get("source_url"),
                "img_width"    : r["image_size"]["width"],
                "img_height"   : r["image_size"]["height"],
                "detected_at"  : r["detected_at"],
                "class_id"     : None,
                "class_name"   : "no_detection",
                "class_en"     : "no_detection",
                "confidence"   : 0.0,
                "x1": None, "y1": None, "x2": None, "y2": None,
                "obj_width"    : None,
                "obj_height"   : None,
            })
        for det in r["detections"]:
            rows.append({
                "image_name"   : r["image_name"],
                "source_url"   : r.get("source_url"),
                "img_width"    : r["image_size"]["width"],
                "img_height"   : r["image_size"]["height"],
                "detected_at"  : r["detected_at"],
                "class_id"     : det["class_id"],
                "class_name"   : det["class_name"],
                "class_en"     : det["class_en"],
                "confidence"   : det["confidence"],
                "x1"           : det["bbox"]["x1"],
                "y1"           : det["bbox"]["y1"],
                "x2"           : det["bbox"]["x2"],
                "y2"           : det["bbox"]["y2"],
                "obj_width"    : det["size_px"]["width"],
                "obj_height"   : det["size_px"]["height"],
            })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────
#  Excel
# ─────────────────────────────────────────────

def _style_sheet(ws, is_rtl: bool = True):
    ws.sheet_view.rightToLeft = is_rtl
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="EEF2FF")
    h_align     = "right" if is_rtl else "left"

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal=h_align, vertical="center")

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"


def save_excel(results: list[dict], path: Path, is_rtl: bool = True) -> Path:
    df       = results_to_df(results)
    detected = df[df["class_name"] != "no_detection"]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        df.to_excel(writer, sheet_name="Raw Detections", index=False)

        if not detected.empty:
            summary = (
                detected.groupby("class_name")
                .agg(
                    total      = ("class_name", "count"),
                    images     = ("image_name", "nunique"),
                    avg_conf   = ("confidence", "mean"),
                    min_conf   = ("confidence", "min"),
                    max_conf   = ("confidence", "max"),
                    avg_width  = ("obj_width",  "mean"),
                    avg_height = ("obj_height", "mean"),
                )
                .round(4).reset_index()
                .sort_values("total", ascending=False)
            )
            summary.to_excel(writer, sheet_name="Class Summary", index=False)

            img_summary = (
                detected.groupby("image_name")
                .agg(
                    total_objects  = ("class_name", "count"),
                    unique_classes = ("class_name", "nunique"),
                    classes_found  = ("class_name", lambda x: ", ".join(sorted(x.unique()))),
                    avg_confidence = ("confidence", "mean"),
                )
                .round(4).reset_index()
                .sort_values("total_objects", ascending=False)
            )
            img_summary.to_excel(writer, sheet_name="Per Image Summary", index=False)

        stats = {
            "Metric": [
                "Total Images", "Images With Detections",
                "Total Objects", "Unique Classes",
                "Avg Confidence", "Generated At",
            ],
            "Value": [
                df["image_name"].nunique(),
                detected["image_name"].nunique() if not detected.empty else 0,
                len(detected),
                detected["class_name"].nunique() if not detected.empty else 0,
                f"{detected['confidence'].mean():.2%}" if not detected.empty else "N/A",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        }
        pd.DataFrame(stats).to_excel(writer, sheet_name="Stats Overview", index=False)

        for name in writer.sheets:
            _style_sheet(writer.sheets[name], is_rtl=is_rtl)

    return path


def save_csv(results: list[dict], path: Path) -> Path:
    results_to_df(results).to_csv(path, index=False, encoding="utf-8-sig")
    return path
